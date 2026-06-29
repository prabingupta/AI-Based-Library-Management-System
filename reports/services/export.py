from io import BytesIO
from datetime import datetime


def generate_books_pdf():
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate,
        Table,
        TableStyle,
        Paragraph,
        Spacer,
        HRFlowable,
    )
    from books.models import Book

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=1.5 * cm,
        leftMargin=1.5 * cm,
        topMargin=2 * cm,
        bottomMargin=1.5 * cm,
    )

    styles = getSampleStyleSheet()
    primary = colors.HexColor("#4f46e5")
    light_gray = colors.HexColor("#f9fafb")
    border_color = colors.HexColor("#e5e7eb")

    title_style = ParagraphStyle(
        "Title",
        parent=styles["Title"],
        fontSize=22,
        textColor=primary,
        spaceAfter=4,
    )
    subtitle_style = ParagraphStyle(
        "Sub",
        parent=styles["Normal"],
        fontSize=10,
        textColor=colors.HexColor("#6b7280"),
        spaceAfter=16,
    )
    heading_style = ParagraphStyle(
        "Heading",
        parent=styles["Normal"],
        fontSize=13,
        textColor=primary,
        fontName="Helvetica-Bold",
        spaceAfter=8,
    )

    elements = []

    elements.append(Paragraph("LibraryOS", title_style))
    elements.append(
        Paragraph(
            f"Books Report — Generated {datetime.now().strftime('%B %d, %Y at %I:%M %p')}",
            subtitle_style,
        )
    )
    elements.append(HRFlowable(width="100%", thickness=1, color=primary))
    elements.append(Spacer(1, 16))

    books = (
        Book.objects.filter(is_deleted=False)
        .prefetch_related("authors", "categories")
        .select_related("publisher", "language")
        .order_by("title")
    )

    elements.append(Paragraph(f"All Books ({books.count()})", heading_style))

    data = [["Title", "Author(s)", "ISBN", "Status", "Copies", "Available"]]
    for book in books:
        authors = ", ".join(a.full_name for a in book.authors.all()) or "—"
        data.append(
            [
                Paragraph(book.title[:45], styles["Normal"]),
                Paragraph(authors[:30], styles["Normal"]),
                book.isbn or book.isbn13 or "—",
                book.get_status_display(),
                str(book.total_copies),
                str(book.available_copies),
            ]
        )

    table = Table(
        data,
        colWidths=[5.5 * cm, 4 * cm, 3.2 * cm, 2.5 * cm, 1.6 * cm, 1.8 * cm],
        repeatRows=1,
    )
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), primary),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 9),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, light_gray]),
                ("FONTSIZE", (0, 1), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.5, border_color),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    elements.append(table)
    elements.append(Spacer(1, 20))

    elements.append(HRFlowable(width="100%", thickness=0.5, color=border_color))
    elements.append(Spacer(1, 6))
    elements.append(
        Paragraph(
            "LibraryOS — AI-Based Library Management System",
            ParagraphStyle("Footer", parent=styles["Normal"], fontSize=8, textColor=colors.HexColor("#9ca3af")),
        )
    )

    doc.build(elements)
    buffer.seek(0)
    return buffer


def generate_members_pdf():
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate,
        Table,
        TableStyle,
        Paragraph,
        Spacer,
        HRFlowable,
    )
    from members.models import Member

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=1.5 * cm,
        leftMargin=1.5 * cm,
        topMargin=2 * cm,
        bottomMargin=1.5 * cm,
    )

    styles = getSampleStyleSheet()
    primary = colors.HexColor("#4f46e5")
    light_gray = colors.HexColor("#f9fafb")
    border_color = colors.HexColor("#e5e7eb")

    title_style = ParagraphStyle(
        "Title",
        parent=styles["Title"],
        fontSize=22,
        textColor=primary,
        spaceAfter=4,
    )
    subtitle_style = ParagraphStyle(
        "Sub",
        parent=styles["Normal"],
        fontSize=10,
        textColor=colors.HexColor("#6b7280"),
        spaceAfter=16,
    )

    elements = []
    elements.append(Paragraph("LibraryOS", title_style))
    elements.append(
        Paragraph(
            f"Members Report — Generated {datetime.now().strftime('%B %d, %Y at %I:%M %p')}",
            subtitle_style,
        )
    )
    elements.append(HRFlowable(width="100%", thickness=1, color=primary))
    elements.append(Spacer(1, 16))

    members = Member.objects.filter(is_deleted=False).select_related("user").order_by("user__last_name")

    data = [["Name", "Card No.", "Type", "Status", "Fine (Rs.)", "Borrows"]]
    for m in members:
        data.append(
            [
                m.user.get_full_name() or m.user.email,
                m.card_number,
                m.get_membership_type_display(),
                m.get_status_display(),
                str(m.fine_balance),
                str(m.current_borrow_count),
            ]
        )

    table = Table(
        data,
        colWidths=[5 * cm, 3.5 * cm, 3 * cm, 2.5 * cm, 2.5 * cm, 2 * cm],
        repeatRows=1,
    )
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), primary),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 9),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, light_gray]),
                ("FONTSIZE", (0, 1), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.5, border_color),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)
    return buffer


def generate_books_excel():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from books.models import Book

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Books Report"

    primary_fill = PatternFill("solid", fgColor="4f46e5")
    alt_fill = PatternFill("solid", fgColor="f9fafb")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    thin_border = Border(
        left=Side(style="thin", color="e5e7eb"),
        right=Side(style="thin", color="e5e7eb"),
        top=Side(style="thin", color="e5e7eb"),
        bottom=Side(style="thin", color="e5e7eb"),
    )

    ws["A1"] = "LibraryOS — Books Report"
    ws["A1"].font = Font(bold=True, size=16, color="4f46e5")
    ws["A2"] = f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
    ws["A2"].font = Font(size=9, color="6b7280")
    ws.append([])

    headers = [
        "Title",
        "Author(s)",
        "Publisher",
        "ISBN",
        "ISBN-13",
        "Language",
        "Pages",
        "Year",
        "Status",
        "Total Copies",
        "Available",
    ]
    ws.append(headers)

    header_row = ws.max_row
    for col, _ in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col)
        cell.fill = primary_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    books = (
        Book.objects.filter(is_deleted=False)
        .prefetch_related("authors")
        .select_related("publisher", "language")
        .order_by("title")
    )

    for i, book in enumerate(books):
        authors = ", ".join(a.full_name for a in book.authors.all()) or "—"
        row = [
            book.title,
            authors,
            book.publisher.name if book.publisher else "—",
            book.isbn or "—",
            book.isbn13 or "—",
            book.language.name if book.language else "—",
            book.pages or "—",
            book.publication_year or "—",
            book.get_status_display(),
            book.total_copies,
            book.available_copies,
        ]
        ws.append(row)
        data_row = ws.max_row
        fill = alt_fill if i % 2 == 1 else PatternFill("solid", fgColor="FFFFFF")
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=data_row, column=col)
            cell.fill = fill
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    col_widths = [40, 25, 20, 15, 16, 12, 8, 8, 14, 14, 12]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    ws.row_dimensions[header_row].height = 22
    ws.freeze_panes = f"A{header_row + 1}"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generate_borrowing_excel():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from borrowing.models import BorrowRecord

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Borrowing Report"

    primary_fill = PatternFill("solid", fgColor="4f46e5")
    alt_fill = PatternFill("solid", fgColor="f9fafb")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    thin_border = Border(
        left=Side(style="thin", color="e5e7eb"),
        right=Side(style="thin", color="e5e7eb"),
        top=Side(style="thin", color="e5e7eb"),
        bottom=Side(style="thin", color="e5e7eb"),
    )

    ws["A1"] = "LibraryOS — Borrowing Report"
    ws["A1"].font = Font(bold=True, size=16, color="4f46e5")
    ws["A2"] = f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
    ws["A2"].font = Font(size=9, color="6b7280")
    ws.append([])

    headers = [
        "Member",
        "Card No.",
        "Book Title",
        "Barcode",
        "Borrow Date",
        "Due Date",
        "Return Date",
        "Status",
        "Fine (Rs.)",
        "Fine Paid",
    ]
    ws.append(headers)

    header_row = ws.max_row
    for col, _ in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col)
        cell.fill = primary_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    records = BorrowRecord.objects.select_related("member__user", "book_copy__book").order_by("-borrow_date")

    for i, record in enumerate(records):
        row = [
            record.member.user.get_full_name(),
            record.member.card_number,
            record.book_copy.book.title,
            record.book_copy.barcode,
            record.borrow_date.strftime("%Y-%m-%d") if record.borrow_date else "—",
            record.due_date.strftime("%Y-%m-%d") if record.due_date else "—",
            record.return_date.strftime("%Y-%m-%d") if record.return_date else "—",
            record.get_status_display(),
            float(record.fine_amount),
            "Yes" if record.fine_paid else "No",
        ]
        ws.append(row)
        data_row = ws.max_row
        fill = alt_fill if i % 2 == 1 else PatternFill("solid", fgColor="FFFFFF")
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=data_row, column=col)
            cell.fill = fill
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    col_widths = [22, 12, 35, 14, 13, 13, 13, 12, 12, 10]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    ws.row_dimensions[header_row].height = 22
    ws.freeze_panes = f"A{header_row + 1}"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
